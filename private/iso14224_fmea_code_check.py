"""
ISO 14224 FMEA failure-mode code compliance checker

Purpose
-------
Checks an Excel FMEA worksheet for rows where the ISO 14224 failure mode (FM) code is:
  1. missing,
  2. not in the ISO 14224 Annex B failure-mode code list, or
  3. a valid ISO 14224 code, but not allowed for the row's equipment class.

This script is tailored to the column names in AutoclaveControlLoopFMEA.xlsx, but the
header lookup is case/space tolerant enough to handle minor edits.

Install
-------
    pip install openpyxl

Run
---
    python iso14224_fmea_code_check.py AutoclaveControlLoopFMEA.xlsx AutoclaveControlLoopFMEA_ISO14224_check.xlsx
"""

from __future__ import annotations

import re
import sys
from copy import copy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# ISO 14224 Annex B failure-mode code list from Tables B.6 to B.12
# -----------------------------------------------------------------------------
# This is the union of failure mode codes appearing across the Annex B equipment
# failure-mode tables. The second validation step then checks whether a valid code
# is allowed for the row's specific equipment class.
ISO14224_ALL_FM_CODES = {
    # common rotating/mechanical/electrical/safety/subsea/drilling codes
    "AIR", "BRD", "BRO", "CLW", "DOP", "ELF", "ELP", "ELU", "ERO",
    "FCO", "FDC", "FOF", "FOV", "FTC", "FTD", "FTF", "FTI", "FTL", "FTO",
    "FTS", "HIO", "IHT", "INL", "LBP", "LCP", "LOA", "LOB", "LOO", "LOR",
    "MOF", "NOI", "NOO", "NON", "OHE", "OTH", "PCL", "PDE", "PLU", "POD",
    "POW", "PTF", "SER", "SET", "SHH", "SLL", "SLP", "SPO", "STD", "STP",
    "UNK", "VIB", "VLO", "WGL",
}

# ISO 14224 Table B.9 - Safety and control equipment failure-mode codes.
# The user requested both checking against the ISO FM code list and against
# the equipment-class columns in the ISO table. For this valve/control-loop FMEA,
# the relevant equipment classes are the Table B.9 classes below.
ISO14224_B9_ALLOWED_BY_CLASS = {
    "Fire detectors": {"FTF", "SPO", "HIO", "LOO", "ERO", "NOO", "SHH", "SLL", "SER", "OTH", "UNK"},
    "Gas detectors": {"SPO", "HIO", "LOO", "VLO", "SHH", "SLL", "SER", "OTH", "UNK"},
    "Input devices": {"FTF", "SPO", "HIO", "LOO", "ERO", "NOO", "SER", "OTH", "UNK"},
    "Control logic units": {"FTF", "SPO", "HIO", "LOO", "ERO", "SER", "OTH", "UNK"},
    "Valves": {"FTO", "FTC", "DOP", "SPO", "HIO", "LOO", "ERO", "PLU", "ELP", "ELU", "INL", "LCP", "AIR", "STD", "OTH", "UNK"},
}

# Short labels used in the FMEA sheet.
EQUIPMENT_CLASS_CODE_MAP = {
    "FD": "Fire detectors",
    "GD": "Gas detectors",
    "ID": "Input devices",
    "CL": "Control logic units",
    "VA": "Valves",
}


# -----------------------------------------------------------------------------
# Utility functions
# -----------------------------------------------------------------------------
# this function is used to normalize both header names and code values for tolerant lookup and comparison.
def norm_header(value: Any) -> str:
    """Normalize a spreadsheet header for tolerant lookup."""
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())

# This function is used to normalize FM code values for comparison against the ISO code list and allowed codes by class.
def norm_code(value: Any) -> str:
    """Normalize an ISO code value."""
    if value is None:
        return ""
    return str(value).strip().upper()

# this function is used to build a mapping of normalized header names to their column indices for flexible column lookup.
# A column indices is returned as 1-based to match openpyxl's cell access patterns. What is a cell access pattern? For example, ws.cell(row=1, column=3) to access the header of the 3rd column.
def build_header_index(ws) -> Dict[str, int]:
    """Return normalized header -> 1-based column index."""
    return {norm_header(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}

#  this function is used to find the column index for a given header name by trying multiple candidate header strings, which allows for minor variations in the header names. For example, the script can find the column for "Component information" even if the header is slightly different like "Component information_added".
def find_col(headers: Dict[str, int], *candidate_headers: str) -> Optional[int]:
    """Find a column by trying several possible header strings."""
    for candidate in candidate_headers:
        key = norm_header(candidate)
        if key in headers:
            return headers[key]
    return None

# this function is used to safely retrieve the value of a cell given its row and column indices, returning None if the column index is not found. This helps to avoid errors when trying to access a cell in a column that may not exist in the input spreadsheet.
def cell_value(ws, row: int, col: Optional[int]) -> Any:
    if not col:
        return None
    return ws.cell(row, col).value

# this function is used to infer the equipment class for a given FMEA row based on the component ID and component information. It applies a set of rules based on common keywords in the component information and common tag prefixes in the component ID to determine the most likely equipment class when it is not explicitly provided in the FMEA row. This inference is important for validating whether the assigned FM code is appropriate for the equipment class according to ISO 14224 Table B.9.
# an example of a rule set is that if the component information contains keywords like "valve" and the component ID does not start with typical input device prefixes, then the equipment class is inferred as "Valves". If the component information contains keywords like "sensor", "transmitter", "switch", "indicator", or "input", then the equipment class is inferred as "Input devices". If the component information contains keywords like "logic", "controller", "plc", or "control", then the equipment class is inferred as "Control logic units". Additionally, if the component ID starts with common instrument tag prefixes used in process control loops (e.g., "LS", "TE", "TI" for input devices, or "XV", "LV" for valves), then the equipment class is inferred accordingly. If none of these rules apply, the function returns None, indicating that the equipment class could not be inferred.
def infer_equipment_class(component_id: Any, component_info: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    Infer ISO 14224 Table B.9 equipment class where the FMEA row does not explicitly
    provide one. This is a conservative rule set for control-loop FMEA rows.
    """
    comp_id = str(component_id or "").upper()
    info = str(component_info or "").lower()

    # Text-based rules first.
    if any(word in info for word in ["valve", "actuated on/off valve", "on/off valve"]):
        # Position switch IDs are usually input devices even when component text says valve.
        if comp_id.startswith(("ZSO", "ZSC", "ZI")):
            return "Input devices", "inferred from component ID"
        return "Valves", "inferred from component information"

    if any(word in info for word in ["sensor", "transmitter", "switch", "indicator", "input"]):
        return "Input devices", "inferred from component information"

    if any(word in info for word in ["logic", "controller", "plc", "control"]):
        return "Control logic units", "inferred from component information"

    # Instrument tag prefixes commonly used in process control loops.
    if comp_id.startswith(("LS", "LSH", "TE", "TT", "TI", "PIT", "PT", "FIT", "FT", "ZI", "ZSO", "ZSC")):
        return "Input devices", "inferred from component ID"
    if comp_id.startswith(("XV", "LV", "PV", "FV", "TV")):
        return "Valves", "inferred from component ID"

    return None, None

# this function is used to generate a practical remediation suggestion for each flagged row based on the specific compliance issue identified. The suggestions are tailored to guide the user on how to correct the issue, such as selecting a valid ISO 14224 FM code, replacing an invalid code with a valid one, or reviewing the equipment class and choosing an appropriate failure-mode code allowed for that class. The function takes into account the nature of the issue, the equipment class (if determined), and the FM code in question to provide actionable advice for resolving the compliance issues in the FMEA worksheet.
def suggest_action(issue: str, equipment_class: Optional[str], fm_code: str) -> str:
    """Create a practical remediation suggestion for each flagged row."""
    if "Missing" in issue:
        return "Select the closest ISO 14224 FM code from the relevant equipment-class table."
    if "not in ISO" in issue:
        if fm_code == "PLO":
            return "Possible typo for PLU (Plugged/choked), but only use PLU where the equipment-class table allows it."
        return "Replace with a valid ISO 14224 Annex B failure-mode code."
    if equipment_class == "Valves" and fm_code == "FTF":
        return "For valves, use FTO, FTC, DOP, SPO, PLU, INL, LCP, etc. as applicable; FTF is not shown for valves in Table B.9."
    if equipment_class == "Control logic units" and fm_code == "AIR":
        return "AIR is not shown for control logic units in Table B.9; check the equipment class or choose an allowed control-logic failure mode."
    if equipment_class == "Input devices" and fm_code == "AIR":
        return "AIR is shown for valves in Table B.9, not input devices; consider HIO, LOO, ERO, NOO or another allowed input-device FM code."
    if equipment_class == "Input devices" and fm_code == "PLU":
        return "For sensors/transmitters/switches, blocked tapping/ingress is better represented as mechanism/cause; use an allowed input-device FM such as FTF, HIO, LOO, ERO, NOO, etc. as applicable."
    return "Review equipment class and choose a failure-mode code allowed for that class."


# -----------------------------------------------------------------------------
# Main checking logic
# -----------------------------------------------------------------------------
# this function is the main entry point for the script, which takes an input Excel file path and an output Excel file path. It loads the input workbook, identifies the relevant columns based on the headers, and iterates through each row of the FMEA worksheet to check the ISO 14224 failure mode code compliance according to the defined rules. For each row, it normalizes the FM code and equipment class information, determines if there are any compliance issues, and collects flagged rows with details about the issues and suggested actions. Finally, it generates new sheets in the workbook for the compliance check results, a summary of issues, and a reference of allowed FM codes by equipment class before saving the output workbook.

def check_fmea(input_xlsx: Path, output_xlsx: Path) -> None:
    wb = load_workbook(input_xlsx)
    ws = wb.active

    headers = build_header_index(ws)

    col_asset_id = find_col(headers, "AssetID")
    col_functional_location = find_col(headers, "FunctionalLocation")
    col_component_id = find_col(headers, "Component_ID")
    col_component_info = find_col(headers, "Component information_added", "Component information")
    col_eq_class_code = find_col(headers, "ISO14224 Equipment Class_added", "ISO14224 Equipment Class")
    col_eq_class_desc = find_col(headers, "ISO14224 Equipment ClassDescription_added", "ISO14224 Equipment Class Description")
    col_fm_code = find_col(headers, "ISO14224_FailureModeCode_added", "ISO14224 FailureModeCode")
    col_fm_desc = find_col(headers, "ISO14224_FailureModeDescription_added", "ISO14224 FailureModeDescription")
    col_company_fm = find_col(headers, "Company_FailureMode don't add", "Company FailureMode")
    col_company_mech = find_col(headers, "Company_FailureMechanism don't add", "Company FailureMechanism")

    required = {
        "ISO14224_FailureModeCode_added": col_fm_code,
        "Component_ID": col_component_id,
        "Component information_added": col_component_info,
    }
    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(f"Missing required input columns: {', '.join(missing)}")

    flagged_rows: List[List[Any]] = []

    for row in range(2, ws.max_row + 1):
        raw_code = cell_value(ws, row, col_fm_code)
        fm_code = norm_code(raw_code)
        eq_code = norm_code(cell_value(ws, row, col_eq_class_code))
        eq_desc = cell_value(ws, row, col_eq_class_desc)
        component_id = cell_value(ws, row, col_component_id)
        component_info = cell_value(ws, row, col_component_info)

        # Determine class: prefer explicit class code/description, otherwise infer.
        if eq_code in EQUIPMENT_CLASS_CODE_MAP:
            checked_class = EQUIPMENT_CLASS_CODE_MAP[eq_code]
            class_source = "explicit"
        elif eq_desc and str(eq_desc).strip() in ISO14224_B9_ALLOWED_BY_CLASS:
            checked_class = str(eq_desc).strip()
            class_source = "explicit"
        else:
            checked_class, class_source = infer_equipment_class(component_id, component_info)

        issue = None
        if not fm_code:
            issue = "Missing ISO 14224 failure mode code"
        elif fm_code not in ISO14224_ALL_FM_CODES:
            issue = "FM code is not in ISO 14224 Annex B failure mode code list"
        elif checked_class in ISO14224_B9_ALLOWED_BY_CLASS:
            allowed = ISO14224_B9_ALLOWED_BY_CLASS[checked_class]
            if fm_code not in allowed:
                issue = f"FM code is valid in ISO 14224 but not shown for equipment class '{checked_class}' in Table B.9"
        elif checked_class is None:
            issue = "Unable to determine ISO 14224 equipment class for class-specific validation"

        if issue:
            flagged_rows.append([
                row,
                cell_value(ws, row, col_asset_id),
                cell_value(ws, row, col_functional_location),
                component_id,
                component_info,
                eq_code or None,
                eq_desc,
                checked_class,
                class_source,
                fm_code or None,
                cell_value(ws, row, col_fm_desc),
                cell_value(ws, row, col_company_fm),
                cell_value(ws, row, col_company_mech),
                issue,
                suggest_action(issue, checked_class, fm_code),
            ])

    # Remove old generated sheets if script is re-run.
    for sheet_name in ["ISO14224_Check", "ISO14224_Summary", "ISO14224_Ref_B9"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    write_check_sheet(wb, flagged_rows)
    write_summary_sheet(wb, flagged_rows)
    write_reference_sheet(wb)

    wb.save(output_xlsx)

# this function is used to create a new sheet in the workbook that lists all the rows from the FMEA worksheet that were flagged for ISO 14224 FM code compliance issues. The sheet includes detailed information about each flagged row, such as the asset ID, functional location, component ID and information, equipment class code and description, the checked or inferred equipment class, the FM code and description, company failure mode and mechanism, the specific compliance issue identified, and a suggested action for remediation. The function also applies styling to the header row and highlights the compliance issue column for better visibility.
def write_check_sheet(wb, flagged_rows: List[List[Any]]) -> None:
    ws = wb.create_sheet("ISO14224_Check")
    headers = [
        "Excel Row", "AssetID", "FunctionalLocation", "Component_ID", "Component information",
        "FMEA Equipment Class Code", "FMEA Equipment Class Description", "Checked/Inferred Class",
        "Class Source", "FMEA FM Code", "FMEA FM Description", "Company Failure Mode",
        "Company Failure Mechanism", "Compliance Issue", "Suggested Action",
    ]
    ws.append(headers)
    for row in flagged_rows:
        ws.append(row)

    style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws, max_width=55)

    # Highlight issue column.
    issue_col = headers.index("Compliance Issue") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row, issue_col).fill = PatternFill("solid", fgColor="FCE4D6")

# this function is used to create a summary sheet in the workbook that provides an overview of the ISO 14224 FM code compliance issues identified in the FMEA worksheet. The summary includes the total number of flagged rows, a breakdown of compliance issues by type with counts for each issue, and a basis of check section that explains the criteria used for the compliance checks. The function also applies styling to the headers and formats the columns for better readability.
def write_summary_sheet(wb, flagged_rows: List[List[Any]]) -> None:
    ws = wb.create_sheet("ISO14224_Summary")
    ws["A1"] = "ISO 14224 FMEA code compliance summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Total flagged rows"
    ws["B3"] = len(flagged_rows)

    counts: Dict[str, int] = {}
    for row in flagged_rows:
        issue = row[13]
        counts[issue] = counts.get(issue, 0) + 1

    ws["A5"] = "Compliance issue"
    ws["B5"] = "Count"
    style_header(ws, 2, row=5)
    for issue, count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        ws.append([issue, count])

    ws["A12"] = "Basis of check"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "1. FM code is present in ISO 14224 Annex B Tables B.6-B.12."
    ws["A14"] = "2. For the valve/control-loop FMEA, the code is allowed for the relevant Table B.9 equipment class."
    autosize(ws, max_width=85)

# this function is used to create a reference sheet in the workbook that lists the allowed ISO 14224 FM codes for each equipment class according to Table B.9 of the standard. The sheet includes columns for the equipment class, the allowed FM codes, the source table (Table B.9), and notes about the types of equipment covered. This reference sheet serves as a quick guide for users to understand which FM codes are valid for each equipment class when reviewing and correcting compliance issues in the FMEA worksheet.
def write_reference_sheet(wb) -> None:
    ws = wb.create_sheet("ISO14224_Ref_B9")
    ws.append(["Equipment class", "Allowed FM codes from ISO 14224 Table B.9", "Source table", "Notes"])
    for equipment_class, codes in ISO14224_B9_ALLOWED_BY_CLASS.items():
        ws.append([equipment_class, ", ".join(sorted(codes)), "ISO 14224 Table B.9", "Safety and control equipment"])
    style_header(ws, 4)
    autosize(ws, max_width=80)

# this function is used to apply consistent styling to the header row of the generated sheets, including a solid fill color, white bold font, and wrapped text alignment. The function takes the worksheet, the maximum number of columns to style, and an optional row number for the header (defaulting to 1) as parameters. This styling helps to visually distinguish the header row from the data rows and improves readability of the generated sheets.
def style_header(ws, max_col: int, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# this function is used to automatically adjust the width of columns in the generated sheets based on the maximum length of the content in each column, up to a specified maximum width. The function iterates through each column and row to determine the appropriate width for each column, ensuring that the content is displayed clearly without excessive whitespace. Additionally, it applies wrapped text alignment to all cells in the column to improve readability when content exceeds the column width.
def autosize(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) + 2))
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[letter].width = width

# this block is the entry point of the script when run from the command line. It checks that exactly two command-line arguments are provided (the input and output Excel file paths), and if not, it prints usage instructions and exits. If the correct arguments are provided, it calls the check_fmea function with the input and output file paths to perform the compliance check and generate the output workbook.
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python iso14224_fmea_code_check.py INPUT.xlsx OUTPUT.xlsx")
        sys.exit(2)
    check_fmea(Path(sys.argv[1]), Path(sys.argv[2]))
