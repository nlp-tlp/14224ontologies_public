"""
ISO 14224 Annex B FMEA failure-mode code compliance checker.

Usage:
    python iso14224_fmea_compliance_check.py "Copy of INPUT.xlsx" "ISO14224_FMEA_Compliance_Check.xlsx"

Assumptions:
- The FMEA worksheet has columns:
  F: ISO14224 Equipment Class_added
  G: ISO14224 Equipment ClassDescription_added
  H: ISO14224_FailureModeCode_added
- ISO 14224:2006 Annex B Tables B.6-B.12 are encoded below.
- Rows with a valid FM code but missing/unrecognised equipment class are flagged because
  equipment-class applicability cannot be confirmed.
"""
from __future__ import annotations

import os
import sys
from collections import Counter
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ISO_VALID_BY_CLASS = {
    "Combustion engine": set("FTS STP UST BRD HIO LOO ERO ELF ELU INL VIB NOI OHE PDE AIR STD SER OTH UNK".split()),
    "Compressor": set("FTS STP UST BRD HIO LOO ERO ELP ELU INL VIB NOI OHE PLU PDE AIR STD SER OTH UNK".split()),
    "Electric generator": set("FTS UST BRD HIO LOO ERO ELU INL VIB NOI OHE PLU PDE AIR STD SER OTH UNK".split()),
    "Electric motor": set("FTS STP UST BRD HIO LOO ERO ELU INL VIB NOI OHE PDE AIR STD SER OTH UNK".split()),
    "Gas turbine": set("FTS STP UST BRD HIO LOO ERO ELF ELP ELU INL VIB NOI OHE PLU PDE AIR STD SER OTH UNK".split()),
    "Pump": set("FTS UST BRD HIO LOO ERO ELP ELU INL VIB NOI OHE PLU PDE AIR STD SER OTH UNK".split()),
    "Steam turbine": set("FTS STP UST BRD HIO LOO ERO ELP ELU INL VIB NOI OHE PLU PDE AIR STD SER OTH UNK".split()),
    "Turbo expander": set("FTS UST BRD HIO LOO ERO ELP ELU INL VIB NOI OHE PLU PDE AIR STD SER OTH UNK".split()),
    "Cranes": set("AIR BRD FCO FTI FRO FTS STP LOA NOI OHE PTF SLP SPO STD PDE VIB SER OTH UNK".split()),
    "Heat exchangers": set("AIR IHT ELP ELU INL PLU NOI OHE STD PDE SER OTH UNK".split()),
    "Heaters and boilers": set("AIR IHT ELP ELU INL PLU NOI OHE STD PDE SER OTH UNK".split()),
    "Piping": set("AIR ELP ELU INL PLU NOI OHE STD PDE VIB SER OTH UNK".split()),
    "Vessels": set("AIR ELP INL PLU LBP LOO LOB MOF NOI OHE PTF STD PDE VIB SER OTH UNK".split()),
    "Winches": set("AIR BRD FTI FRO FTS STP LOA NOI OHE PTF SLP SPO STD PDE VIB SER OTH UNK".split()),
    "Turrets": set("AIR FCO FTI FDC LOB MOF PTF SPO STD PDE SER OTH UNK".split()),
    "Swivels": set("AIR ELP ELU FCO FTI FDC INL LBP PLU STD PDE SER OTH UNK".split()),
    "UPS": set("FTF FOF FOV LOR ERO OHE PDE SPO SER OTH UNK".split()),
    "Power transformers": set("FTF FOV OHE PDE AIR PLU ELU STD INL SER OTH UNK".split()),
    "Fire detectors": set("FTF SPO HIO LOO ERO NOO SHH SLL SER OTH UNK".split()),
    "Gas detectors": set("SPO HIO LOO VLO NOO SHH SLL SER OTH UNK".split()),
    "Input devices": set("FTF SPO HIO LOO ERO NOO ELP ELU SER OTH UNK".split()),
    "Control logic units": set("FTF SPO HIO LOO ERO SER OTH UNK".split()),
    "Valves": set("FTO FTC DOP SPO HIO LOO PLU ELP ELU INL LCP AIR STD SER OTH UNK".split()),
    "Subsea control systems": set("FTF SPO POW LOR PLU ELP ELU INL AIR NON OTH".split()),
    "Xmas trees": set("FTO FTC FTL SET SPO POW LOB PLU ELP ELU INL STD NON OTH".split()),
    "Subsea pumps": set("FTF SPO HIO LOO ELP ELU INL AIR OTH".split()),
    "Risers": set("ELP ELU INL OTH".split()),
    "DHSV": set("FTO FTC LCP WGL CLW PCL OTH UNK".split()),
    "Top drive": set("AIR ELU ERO FTS STP INL HIO LOO NOI OHE SPO VIB SER OTH UNK".split()),
    "Blowout preventer": set("FTF FTO FTC AIR ELU ERO INL LCP SPO STD LOR POD PLU FCO FTD SER OTH UNK".split()),
}

CLASS_ABBREV_TO_NAME = {
    "CE": "Combustion engine", "CO": "Compressor", "EG": "Electric generator", "EM": "Electric motor",
    "GT": "Gas turbine", "PU": "Pump", "ST": "Steam turbine", "TE": "Turbo expander",
    "CR": "Cranes", "HE": "Heat exchangers", "HB": "Heaters and boilers", "PI": "Piping",
    "VE": "Vessels", "WI": "Winches", "TU": "Turrets", "SW": "Swivels",
    "UPS": "UPS", "PT": "Power transformers",
    "FD": "Fire detectors", "GD": "Gas detectors", "ID": "Input devices", "CL": "Control logic units", "VA": "Valves",
    "SCS": "Subsea control systems", "XT": "Xmas trees", "SP": "Subsea pumps", "RI": "Risers",
    "DHSV": "DHSV", "TD": "Top drive", "BOP": "Blowout preventer",
}
ALL_ISO_CODES = sorted(set().union(*ISO_VALID_BY_CLASS.values()))

def normalize_class(class_code, class_desc):
    cc = (class_code or "").strip().upper()
    cd = (class_desc or "").strip()
    if cc in CLASS_ABBREV_TO_NAME:
        return CLASS_ABBREV_TO_NAME[cc]
    for name in ISO_VALID_BY_CLASS:
        if cd.lower() == name.lower():
            return name
    return None

def classify_row(fm_code, class_code, class_desc):
    code = (fm_code or "").strip().upper()
    class_name = normalize_class(class_code, class_desc)
    if not code:
        return "Not assessed", "", "", class_name or "", "No FMEA failure mode code provided.", "", ""
    if code not in ALL_ISO_CODES:
        suggestion = ", ".join(get_close_matches(code, ALL_ISO_CODES, n=3, cutoff=0.65))
        return "Non-compliant", "No", "No", class_name or "", f"FM code '{code}' is not in the ISO 14224 Annex B failure mode code list.", suggestion, ""
    if not class_name:
        return "Non-compliant", "Yes", "No", "", "FM code is in the ISO list, but the ISO 14224 equipment class is missing or not recognised, so table applicability cannot be confirmed.", "", ""
    valid_codes = ISO_VALID_BY_CLASS[class_name]
    if code not in valid_codes:
        return "Non-compliant", "Yes", "No", class_name, f"FM code '{code}' is not permitted for ISO equipment class '{class_name}' in the ISO 14224 Annex B equipment-class failure mode tables.", "", ", ".join(sorted(valid_codes))
    return "Compliant", "Yes", "Yes", class_name, "FM code is in the ISO list and is permitted for the resolved equipment class.", "", ", ".join(sorted(valid_codes))

def add_table_safe(sheet, name):
    if sheet.max_row > 1 and sheet.max_column > 0:
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(tab)

def run(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    orig_max_col = ws.max_column

    audit_headers = [
        "ISO14224_ComplianceStatus",
        "ISO14224_ValidFMCodeList",
        "ISO14224_ValidForEquipmentClass",
        "ISO14224_ResolvedEquipmentClass",
        "ISO14224_NonComplianceReason",
        "ISO14224_SuggestedCode",
        "ISO14224_ValidCodesForClass",
    ]
    for offset, h in enumerate(audit_headers, start=1):
        cell = ws.cell(row=1, column=orig_max_col + offset, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    noncompliant_records = []
    for row in range(2, ws.max_row + 1):
        result = classify_row(ws.cell(row, 8).value, ws.cell(row, 6).value, ws.cell(row, 7).value)
        for i, val in enumerate(result, start=orig_max_col + 1):
            ws.cell(row=row, column=i, value=val)
            ws.cell(row=row, column=i).alignment = Alignment(wrap_text=True, vertical="top")

        status = result[0]
        fill = PatternFill("solid", fgColor="FFC7CE" if status == "Non-compliant" else "C6EFCE" if status == "Compliant" else "E7E6E6")
        for col in range(orig_max_col + 1, orig_max_col + len(audit_headers) + 1):
            ws.cell(row, col).fill = fill

        if status == "Non-compliant":
            rec = {"Excel_Row": row}
            for c in range(1, orig_max_col + 1):
                rec[headers[c-1] if c-1 < len(headers) and headers[c-1] else f"Column_{c}"] = ws.cell(row, c).value
            for name, val in zip(audit_headers, result):
                rec[name] = val
            noncompliant_records.append(rec)

    if "Noncompliant_Rows" in wb.sheetnames:
        del wb["Noncompliant_Rows"]
    audit_ws = wb.create_sheet("Noncompliant_Rows")
    audit_cols = ["Excel_Row", "AssetID", "AssetType", "FunctionalLocation", "Component_ID",
                  "Component information_added", "ISO14224 Equipment Class_added",
                  "ISO14224 Equipment ClassDescription_added", "ISO14224_FailureModeCode_added",
                  "ISO14224_FailureModeDescription_added", "Company_Functional_Purpose",
                  "Company_FailureMode don't add", "Company_FailureMechanism don't add",
                  "ISO14224_ComplianceStatus", "ISO14224_ValidFMCodeList",
                  "ISO14224_ValidForEquipmentClass", "ISO14224_ResolvedEquipmentClass",
                  "ISO14224_NonComplianceReason", "ISO14224_SuggestedCode", "ISO14224_ValidCodesForClass"]
    audit_ws.append(audit_cols)
    for rec in noncompliant_records:
        audit_ws.append([rec.get(c) for c in audit_cols])

    if "ISO_FM_Code_Reference" in wb.sheetnames:
        del wb["ISO_FM_Code_Reference"]
    ref_ws = wb.create_sheet("ISO_FM_Code_Reference")
    ref_ws.append(["Source", "ISO14224 Equipment Class", "Common Worksheet Abbrev", "Permitted FM Codes"])
    abbr_by_name = {}
    for abbr, name in CLASS_ABBREV_TO_NAME.items():
        abbr_by_name.setdefault(name, []).append(abbr)
    for name, codes in sorted(ISO_VALID_BY_CLASS.items()):
        ref_ws.append(["ISO 14224:2006 Annex B Tables B.6-B.12", name, ", ".join(sorted(abbr_by_name.get(name, []))), ", ".join(sorted(codes))])

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    sum_ws = wb.create_sheet("Summary", 0)
    rows_with_code = sum(1 for r in range(2, ws.max_row + 1) if str(ws.cell(r, 8).value or "").strip())
    sum_ws.append(["ISO 14224 FMEA Failure Mode Code Compliance Check", ""])
    sum_ws.append(["Input workbook", os.path.basename(input_path)])
    sum_ws.append(["Source standard used", "ISO 14224:2006 Annex B, Tables B.6-B.12 failure-mode codes and equipment-class applicability"])
    sum_ws.append(["Rows analysed", ws.max_row - 1])
    sum_ws.append(["Rows with FMEA FM code provided", rows_with_code])
    sum_ws.append(["Non-compliant rows", len(noncompliant_records)])

    breakdown = Counter()
    for rec in noncompliant_records:
        reason = rec["ISO14224_NonComplianceReason"]
        if "not in the ISO" in reason:
            key = "FM code not in ISO 14224 FM code list"
        elif "missing or not recognised" in reason:
            key = "Missing/unrecognised ISO equipment class"
        else:
            key = "FM code not permitted for resolved equipment class"
        breakdown[key] += 1
    sum_ws.append([])
    sum_ws.append(["Non-compliance breakdown", "Count"])
    for key, value in breakdown.items():
        sum_ws.append([key, value])

    for sht in wb.worksheets:
        for row in sht.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in sht[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        sht.freeze_panes = "A2"
        sht.row_dimensions[1].height = 36
        for col in range(1, min(sht.max_column, 35) + 1):
            letter = get_column_letter(col)
            maxlen = 0
            for cell in sht[letter][: min(sht.max_row, 100)]:
                maxlen = max(maxlen, min(len("" if cell.value is None else str(cell.value)), 80))
            sht.column_dimensions[letter].width = max(10, min(maxlen + 2, 45))

    add_table_safe(audit_ws, "NoncompliantRowsTable")
    add_table_safe(ref_ws, "ISOFMCodeReferenceTable")
    wb.save(output_path)
    return len(noncompliant_records)

if __name__ == "__main__":
    in_path = sys.argv[1] if len(sys.argv) > 1 else "Copy of INPUT.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "ISO14224_FMEA_Compliance_Check.xlsx"
    count = run(in_path, out_path)
    print(f"Saved {out_path}; non-compliant rows: {count}")
